import csv
import json
import openpyxl
from openpyxl.styles import PatternFill, Font
from collections import OrderedDict

# ============================================================
# CONFIGURACIÓN
# ============================================================
BILLS_CSV       = "BillLineItem_Bill_Export_2026-03-08_02-16-07.csv"
PAYMENTS_CSV    = "BillPaymentLineItem_BillPayment_Export_2026-03-13_21-18-07.csv"
OUTPUT_XLSX     = "BillPayments_Zoho.xlsx"
OUTPUT_SKIPPED  = "BillPayments_sin_bill.csv"
PAID_THROUGH    = "Cuenta Banco General"   # ajusta si hay otras cuentas
FECHA_DESDE     = "2019-01-01"

ZOHO_COLUMNS = [
    "Payment Number", "Date", "Vendor Name", "Mode",
    "Exchange Rate", "Amount", "Paid Through",
    "Description", "Bill Number", "Bill Amount"
]

# ============================================================
# HELPERS
# ============================================================
def format_date(raw):
    raw = raw.strip()
    if not raw:
        return ""
    if "-" in raw:
        parts = raw.split("-")
        return f"{parts[2]}/{parts[1]}/{parts[0]}"
    parts = raw.split("/")
    if len(parts) == 3:
        return f"{int(parts[1]):02d}/{int(parts[0]):02d}/{parts[2]}"
    return raw

def normalize_date(raw):
    raw = raw.strip()
    if "-" in raw:
        return raw[:10]
    parts = raw.split("/")
    if len(parts) == 3:
        return f"{parts[2]}-{int(parts[0]):02d}-{int(parts[1]):02d}"
    return raw

# ============================================================
# 1. Cargar mapa TxnId → DocNumber desde bills
# ============================================================
print("Cargando mapa de bills QB...")
bill_map = {}  # TxnId(Id) → DocNumber
with open(BILLS_CSV, encoding="utf-8") as f:
    reader = csv.DictReader(f, delimiter=";")
    for row in reader:
        bill_id  = row.get("Id1", "").strip()
        doc      = row.get("DocNumber", "").strip()
        if bill_id:
            bill_map[bill_id] = doc or f"QB-{bill_id}"

print(f"Bills en mapa: {len(bill_map)}")

# ============================================================
# 2. Cargar pagos y agrupar por Id (pago único)
# ============================================================
print("Cargando pagos de proveedores...")
with open(PAYMENTS_CSV, encoding="utf-8") as f:
    reader = csv.DictReader(f, delimiter=";")
    raw_rows = list(reader)

# Agrupar por Id (cada fila es ya un pago individual con un bill vinculado)
print(f"Total filas: {len(raw_rows)}")

# ============================================================
# 3. Transformar
# ============================================================
output_rows = []
skipped     = []
stats       = {"incluidos": 0, "skipped_fecha": 0, "skipped_no_bill": 0, "skipped_journal": 0}

for r in raw_rows:
    txn_date = normalize_date(r.get("TxnDate", ""))
    if txn_date < FECHA_DESDE:
        stats["skipped_fecha"] += 1
        continue

    amount = r.get("Amount", "").strip()
    try:
        if float(amount) == 0:
            continue
    except:
        continue

    # Resolver bill vinculado
    linked_str = r.get("LinkedTxn", "[]")
    try:
        linked = json.loads(linked_str)
    except:
        linked = []

    bill_number = ""
    for t in linked:
        if t.get("TxnType") == "Bill":
            txn_id = str(t.get("TxnId", ""))
            bill_number = bill_map.get(txn_id, "")
            break
        elif t.get("TxnType") == "JournalEntry":
            stats["skipped_journal"] += 1

    if not bill_number:
        skipped.append(r)
        stats["skipped_no_bill"] += 1
        continue

    vendor      = r.get("VendorRefName", "").strip()
    pay_date    = format_date(r.get("TxnDate", ""))
    doc_number  = r.get("DocNumber", "").strip()
    notes       = r.get("PrivateNote", "").strip()
    bank_acct   = r.get("CheckPayment_BankAccountRefName", "").strip() or PAID_THROUGH

    output_rows.append({
        "Payment Number": doc_number,
        "Date":           pay_date,
        "Vendor Name":    vendor,
        "Mode":           "Check",
        "Exchange Rate":  "1",
        "Amount":         amount,
        "Paid Through":   bank_acct,
        "Description":    notes,
        "Bill Number":    bill_number,
        "Bill Amount":    amount,
    })
    stats["incluidos"] += 1

# ============================================================
# 4. Escribir XLSX
# ============================================================
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Bill Payments"

hfill = PatternFill("solid", fgColor="2D6A9F")
hfont = Font(color="FFFFFF", bold=True)
for i, col in enumerate(ZOHO_COLUMNS, 1):
    c = ws.cell(row=1, column=i, value=col)
    c.fill = hfill
    c.font = hfont

for row_data in output_rows:
    ws.append([row_data.get(col, "") for col in ZOHO_COLUMNS])

for col in ws.columns:
    max_len = max((len(str(c.value)) for c in col if c.value), default=10)
    ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 45)

wb.save(OUTPUT_XLSX)

# Skipped log
if skipped:
    with open(OUTPUT_SKIPPED, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=raw_rows[0].keys())
        writer.writeheader()
        writer.writerows(skipped)

print(f"\n{'='*52}")
print(f"✅ Archivo generado      : {OUTPUT_XLSX}")
print(f"   Pagos incluidos       : {stats['incluidos']}")
print(f"   Sin bill (skipped)    : {stats['skipped_no_bill']}")
print(f"   Journal entries skip  : {stats['skipped_journal']}")
print(f"   Anteriores a 2019     : {stats['skipped_fecha']}")
if skipped:
    print(f"⚠️  Log omitidos          : {OUTPUT_SKIPPED}")
print(f"{'='*52}")
