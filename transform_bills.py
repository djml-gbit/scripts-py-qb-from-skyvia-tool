import csv
import json
import openpyxl
from openpyxl.styles import PatternFill, Font
from collections import OrderedDict

# ============================================================
# CONFIGURACIÓN
# ============================================================
INPUT_CSV   = "BillLineItem_Bill_Export_2026-03-08_02-16-07.csv"
OUTPUT_XLSX = "Bills_Zoho.xlsx"
OUTPUT_SKIPPED = "Bills_impuesto_desconocido.csv"
LOTE        = "Todo hasta sep 2025"
FECHA_DESDE = "2019-01-01"

# Tax codes que se procesan normalmente
TAX_INCLUDE = {"9", "3", "5", "12", "NON", ""}  # ITBMS 7% confirmado y exentos
TAX_SKIP    = {"7", "8", "10"}  # ISC/Telecom - revisar con cliente

ZOHO_COLUMNS = [
    "Bill Date", "Bill Number", "Bill Status",
    "Vendor Name", "Due Date", "Currency Code",
    "Account", "Description", "Quantity", "Rate",
    "Tax Name", "Tax Percentage",
    "Vendor Notes", "Lote"
]

# ============================================================
# HELPERS
# ============================================================
def format_date(raw):
    raw = raw.strip()
    if not raw:
        return ""
    if "-" in raw:
        parts = raw.split("-")
        return f"{parts[2]}/{parts[1]}/{parts[0]}"
    parts = raw.split("/")
    if len(parts) == 3:
        return f"{int(parts[1]):02d}/{int(parts[0]):02d}/{parts[2]}"
    return raw

def normalize_date_for_filter(raw):
    """Devuelve YYYY-MM-DD para comparar"""
    raw = raw.strip()
    if not raw:
        return ""
    if "-" in raw:
        return raw[:10]
    parts = raw.split("/")
    if len(parts) == 3:
        return f"{parts[2]}-{int(parts[0]):02d}-{int(parts[1]):02d}"
    return raw

def strip_category(name):
    if ":" in name:
        return name.split(":", 1)[1].strip()
    return name.strip()

def get_tax(tax_id):
    if tax_id in ("9", "3", "5"):   # ITBMS 7% (activo e inactivos confirmados)
        return ("ITBMS DGI", "7")
    return ("", "")                  # 12=DGI EXENTO, NON, vacío = sin impuesto

# ============================================================
# LECTURA Y AGRUPACIÓN
# ============================================================
print("Cargando bills QB...")
with open(INPUT_CSV, encoding="utf-8") as f:
    reader = csv.DictReader(f, delimiter=";")
    raw_rows = list(reader)

# Agrupar por Id1 (ID interno) para manejar DocNumber vacíos
bills = OrderedDict()
for row in raw_rows:
    bill_id = row["Id1"].strip()
    if bill_id not in bills:
        bills[bill_id] = []
    bills[bill_id].append(row)

print(f"Bills totales en QB: {len(bills)}")

# ============================================================
# TRANSFORMACIÓN
# ============================================================
output_rows  = []
skipped_rows = []
stats = {
    "total": 0, "incluidos": 0, "antes_2019": 0,
    "tax_raro": 0, "lineas": 0, "sin_docnum": 0
}

for bill_id, rows in bills.items():
    header = rows[0]
    txn_date = normalize_date_for_filter(header["TxnDate"])

    # Filtrar antes de 2019
    if txn_date < FECHA_DESDE:
        stats["antes_2019"] += 1
        continue

    stats["total"] += 1

    # Verificar si alguna línea tiene tax raro
    has_weird_tax = False
    for r in rows:
        if r["DetailType"] == "AccountBasedExpenseLineDetail":
            tc = r["AccountBasedExpenseLineDetail_TaxCodeRefId"].strip()
            if tc in TAX_SKIP:
                has_weird_tax = True
                break
        elif r["DetailType"] == "ItemBasedExpenseLineDetail":
            tc = r["ItemBasedExpenseLineDetail_TaxCodeRefId"].strip()
            if tc in TAX_SKIP:
                has_weird_tax = True
                break

    if has_weird_tax:
        stats["tax_raro"] += 1
        skipped_rows.append({
            "Bill Number": header["DocNumber"].strip() or f"QB-{bill_id}",
            "Vendor": header["VendorRefName"].strip(),
            "Date": header["TxnDate"],
            "Total": header["TotalAmt"],
            "Motivo": "Tax code desconocido (no es ITBMS 7% ni exento)"
        })
        continue

    # DocNumber: usar el de QB o auto-generar
    doc_number = header["DocNumber"].strip() or f"QB-{bill_id}"
    if not header["DocNumber"].strip():
        stats["sin_docnum"] += 1

    bill_date  = format_date(header["TxnDate"])
    due_date   = format_date(header["DueDate"])
    vendor     = header["VendorRefName"].strip()
    currency   = header.get("CurrencyRefId", "USD").strip() or "USD"
    notes      = header.get("PrivateNote", "").strip()

    bill_written = False

    for r in rows:
        detail = r["DetailType"].strip()
        amount = r["Amount"].strip()
        try:
            if float(amount) == 0:
                continue
        except:
            continue

        if detail == "AccountBasedExpenseLineDetail":
            account  = r["AccountBasedExpenseLineDetail_AccountRefName"].strip()
            desc     = r["Description"].strip()
            tax_id   = r["AccountBasedExpenseLineDetail_TaxCodeRefId"].strip()
            tax_name, tax_pct = get_tax(tax_id)

            output_rows.append({
                "Bill Date":    bill_date,
                "Bill Number":  doc_number,
                "Bill Status":  "Open",
                "Vendor Name":  vendor,
                "Due Date":     due_date,
                "Currency Code": currency,
                "Account":      account,
                "Description":  desc,
                "Quantity":     "1",
                "Rate":         amount,
                "Tax Name":     tax_name,
                "Tax Percentage": tax_pct,
                "Vendor Notes": notes if not bill_written else "",
                "Lote":         LOTE,
            })
            bill_written = True
            stats["lineas"] += 1

        elif detail == "ItemBasedExpenseLineDetail":
            item_name = strip_category(r["ItemBasedExpenseLineDetail_ItemRefName"])
            qty       = r["ItemBasedExpenseLineDetail_Qty"].strip() or "1"
            price     = r["ItemBasedExpenseLineDetail_UnitPrice"].strip() or amount
            tax_id    = r["ItemBasedExpenseLineDetail_TaxCodeRefId"].strip()
            tax_name, tax_pct = get_tax(tax_id)

            output_rows.append({
                "Bill Date":    bill_date,
                "Bill Number":  doc_number,
                "Bill Status":  "Open",
                "Vendor Name":  vendor,
                "Due Date":     due_date,
                "Currency Code": currency,
                "Account":      item_name,   # Zoho usa Account para ItemBased también
                "Description":  r["Description"].strip(),
                "Quantity":     qty,
                "Rate":         price,
                "Tax Name":     tax_name,
                "Tax Percentage": tax_pct,
                "Vendor Notes": notes if not bill_written else "",
                "Lote":         LOTE,
            })
            bill_written = True
            stats["lineas"] += 1

    if bill_written:
        stats["incluidos"] += 1

# ============================================================
# ESCRITURA XLSX PRINCIPAL
# ============================================================
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Bills"

header_fill = PatternFill("solid", fgColor="2D6A9F")
header_font = Font(color="FFFFFF", bold=True)

for col_idx, col_name in enumerate(ZOHO_COLUMNS, 1):
    cell = ws.cell(row=1, column=col_idx, value=col_name)
    cell.fill = header_fill
    cell.font = header_font

for row_data in output_rows:
    ws.append([row_data.get(col, "") for col in ZOHO_COLUMNS])

for col in ws.columns:
    max_len = max((len(str(c.value)) for c in col if c.value), default=10)
    ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 45)

wb.save(OUTPUT_XLSX)

# ============================================================
# ESCRITURA CSV SKIPPED
# ============================================================
if skipped_rows:
    with open(OUTPUT_SKIPPED, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["Bill Number","Vendor","Date","Total","Motivo"])
        writer.writeheader()
        writer.writerows(skipped_rows)

print(f"\n{'='*52}")
print(f"✅ Archivo generado     : {OUTPUT_XLSX}")
print(f"   Bills incluidos      : {stats['incluidos']}")
print(f"   Líneas generadas     : {stats['lineas']}")
print(f"   Sin DocNumber (QB-)  : {stats['sin_docnum']}")
print(f"   Tax raro (separados) : {stats['tax_raro']}")
print(f"   Anteriores a 2019    : {stats['antes_2019']}")
if skipped_rows:
    print(f"⚠️  Revisar con cliente  : {OUTPUT_SKIPPED}")
print(f"{'='*52}")
