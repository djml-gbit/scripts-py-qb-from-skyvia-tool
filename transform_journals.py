import csv
import openpyxl
from openpyxl.styles import PatternFill, Font
from collections import OrderedDict

# ============================================================
# CONFIGURACIÓN
# ============================================================
INPUT_CSV   = "JournalEntryLineItem_JournalEntry_Export_2026-03-14_19-46-25.csv"
OUTPUT_XLSX = "Journals_Zoho.xlsx"
FECHA_DESDE = "2019-01-01"

ZOHO_COLUMNS = [
    "Journal Date", "Reference Number", "Journal Number Prefix", "Journal Number Suffix",
    "Notes", "Journal Type", "Currency",
    "Account", "Description", "Contact Name",
    "Debit", "Credit", "Status"
]

# ============================================================
# HELPERS
# ============================================================
def format_date(raw):
    raw = raw.strip()
    if not raw: return ""
    if "-" in raw:
        p = raw.split("-"); return f"{p[2]}/{p[1]}/{p[0]}"
    p = raw.split("/")
    if len(p) == 3: return f"{int(p[1]):02d}/{int(p[0]):02d}/{p[2]}"
    return raw

def normalize_date(raw):
    raw = raw.strip()
    if "-" in raw: return raw[:10]
    p = raw.split("/")
    if len(p) == 3: return f"{p[2]}-{int(p[0]):02d}-{int(p[1]):02d}"
    return raw

def strip_category(name):
    if ":" in name:
        return name.split(":", 1)[1].strip()
    return name.strip()

# ============================================================
# LECTURA Y AGRUPACIÓN
# ============================================================
print("Cargando diarios QB...")
with open(INPUT_CSV, encoding="utf-8") as f:
    reader = csv.DictReader(f, delimiter=";")
    raw_rows = list(reader)

journals = OrderedDict()
for row in raw_rows:
    jid = row["Id1"].strip()
    if jid not in journals:
        journals[jid] = []
    journals[jid].append(row)

print(f"Diarios únicos: {len(journals)}")

# ============================================================
# TRANSFORMACIÓN
# ============================================================
output_rows = []
stats = {"incluidos": 0, "skipped_fecha": 0, "skipped_empty": 0}

for jid, rows in journals.items():
    header = rows[0]
    txn_date = normalize_date(header.get("TxnDate", ""))

    if txn_date < FECHA_DESDE:
        stats["skipped_fecha"] += 1
        continue

    journal_date = format_date(header.get("TxnDate", ""))
    doc_number   = header.get("DocNumber", "").strip()
    notes        = header.get("PrivateNote", "").strip()
    currency     = header.get("CurrencyRefId", "USD").strip() or "USD"

    journal_written = False

    for r in rows:
        posting = r.get("JournalEntryLineDetail_PostingType", "").strip()
        amount  = r.get("Amount", "").strip()
        account = r.get("JournalEntryLineDetail_AccountRefName", "").strip()

        if not posting or not account or not amount:
            stats["skipped_empty"] += 1
            continue

        try:
            amt_float = float(amount)
            if amt_float == 0:
                stats["skipped_empty"] += 1
                continue
        except:
            stats["skipped_empty"] += 1
            continue

        account  = strip_category(account)
        desc     = r.get("Description", "").strip()
        contact  = r.get("JournalEntryLineDetail_EntityRefName", "").strip()
        debit    = amount if posting == "Debit"  else "0"
        credit   = amount if posting == "Credit" else "0"

        # Suffix debe ser numérico
        if doc_number.isdigit():
            suffix = doc_number
            prefix = "JE-"
        elif doc_number:
            # Extraer solo dígitos del DocNumber (ej: "528 CM" → "528")
            digits = ''.join(c for c in doc_number if c.isdigit())
            suffix = digits or jid
            prefix = "JE-"
        else:
            suffix = jid  # Id interno de QB como sufijo
            prefix = "QB-"
        output_rows.append({
            "Journal Date":          journal_date,
            "Reference Number":      doc_number or suffix,
            "Journal Number Prefix": prefix,
            "Journal Number Suffix": suffix,
            "Notes":                 notes or desc or "Migrado desde QuickBooks",
            "Journal Type":         "both",
            "Currency":             currency,
            "Account":              account,
            "Description":          desc,
            "Contact Name":         contact,
            "Debit":                debit,
            "Credit":               credit,
            "Status":               "published",
        })
        journal_written = True

    if journal_written:
        stats["incluidos"] += 1

# ============================================================
# ESCRITURA XLSX
# ============================================================
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Journals"

hfill = PatternFill("solid", fgColor="2D6A9F")
hfont = Font(color="FFFFFF", bold=True)
for i, col in enumerate(ZOHO_COLUMNS, 1):
    c = ws.cell(row=1, column=i, value=col)
    c.fill = hfill
    c.font = hfont

for row_data in output_rows:
    ws.append([row_data.get(col, "") for col in ZOHO_COLUMNS])

for col in ws.columns:
    max_len = max((len(str(c.value)) for c in col if c.value), default=10)
    ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 45)

wb.save(OUTPUT_XLSX)

print(f"\n{'='*52}")
print(f"✅ Archivo generado    : {OUTPUT_XLSX}")
print(f"   Diarios incluidos  : {stats['incluidos']}")
print(f"   Líneas generadas   : {len(output_rows)}")
print(f"   Anteriores a 2019  : {stats['skipped_fecha']}")
print(f"   Líneas vacías skip : {stats['skipped_empty']}")
print(f"{'='*52}")
